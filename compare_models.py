#!/usr/bin/env python3
"""Run question sets from Excel through multiple LLMs and grade answers with Claude."""

from __future__ import annotations

import argparse
import os
import re
import sys
import time
from datetime import datetime
from typing import Any

from anthropic import Anthropic
from apify_client import ApifyClient
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

APIFY_ACTOR_ID = "onescales/ai-model-comparison"
ANTHROPIC_EVALUATOR_MODEL = "claude-opus-4-8"
PADDING_MODEL = "OpenAI GPT 5 Mini"
MAX_PROMPT_CHARS = 800
MAX_MODELS_PER_RUN = 4
MIN_MODELS_PER_RUN = 2

RESULT_COLUMNS = [
    "timestamp",
    "question",
    "answer_expected",
    "service_provider",
    "model",
    "answer",
    "grade",
    "comment",
]

DEFAULT_MODELS = [
    ("openai", "OpenAI GPT 5.5"),
    ("openai", "OpenAI GPT 5 Mini"),
    ("anthropic", "Claude Opus 4.7"),
    ("anthropic", "Claude Sonnet 4.5"),
    ("gemini", "Gemini 3.1 Pro Preview"),
    ("gemini", "Gemini 3.0 Pro Preview"),
    ("grok", "Grok 4.3"),  # only one Grok model available on Apify
]

SAMPLE_QUESTIONS = [
    (
        "What is the capital of France?",
        "Paris",
    ),
    (
        "How many continents are there on Earth?",
        "7",
    ),
    (
        "What planet is known as the Red Planet?",
        "Mars",
    ),
]


def model_to_key(model_name: str) -> str:
    return model_name.replace(" ", "_")


def chunk_list(items: list[str], size: int = MAX_MODELS_PER_RUN) -> list[list[str]]:
    unique_items = list(dict.fromkeys(items))
    chunks: list[list[str]] = []
    i = 0
    while i < len(unique_items):
        remaining = len(unique_items) - i
        if remaining <= size:
            chunk = unique_items[i:]
            if len(chunk) == 1:
                chunk = [chunk[0], PADDING_MODEL]
            chunks.append(chunk)
            break
        chunks.append(unique_items[i : i + size])
        i += size
    return chunks


def build_evaluation_prompt(question: str, expected: str, answer: str) -> str:
    return (
        "Grade this AI answer against the expected answer. "
        "The expected answer describes what a good response should contain; "
        "partial credit is fine when key points are present. "
        "Reply exactly on two lines: GRADE: <0-10>/10  COMMENT: <one sentence why>\n"
        f"Q: {question}\nExpected: {expected}\nAnswer: {answer}"
    )


def parse_grade_and_comment(text: str) -> tuple[str, str]:
    grade_match = re.search(r"GRADE:\s*(\d+(?:\.\d+)?)\s*/\s*10", text, re.IGNORECASE)
    comment_match = re.search(r"COMMENT:\s*(.+)", text, re.IGNORECASE | re.DOTALL)
    grade = grade_match.group(1) if grade_match else ""
    comment = comment_match.group(1).strip() if comment_match else text.strip()
    return grade, comment


class ModelComparisonRunner:
    def __init__(self, token: str) -> None:
        self.client = ApifyClient(token)

    def run_actor(self, prompt: str, ai_models: list[str], best_response: str = "") -> dict[str, Any]:
        if len(prompt) > MAX_PROMPT_CHARS:
            raise ValueError(f"Prompt exceeds {MAX_PROMPT_CHARS} characters: {len(prompt)}")
        if not MIN_MODELS_PER_RUN <= len(ai_models) <= MAX_MODELS_PER_RUN:
            raise ValueError(f"ai_models must contain {MIN_MODELS_PER_RUN}-{MAX_MODELS_PER_RUN} items")

        run_input = {
            "prompt": prompt,
            "aiModels": ai_models,
            "bestResponse": best_response,
        }
        run = self.client.actor(APIFY_ACTOR_ID).call(run_input=run_input)
        dataset_id = run.default_dataset_id if hasattr(run, "default_dataset_id") else run["defaultDatasetId"]
        items = list(self.client.dataset(dataset_id).iterate_items())
        if not items:
            raise RuntimeError("Apify actor returned no dataset items")
        return items[0]

    def get_model_answers(self, question: str, models: list[str]) -> dict[str, str]:
        answers: dict[str, str] = {}
        unique_models = list(dict.fromkeys(models))
        for batch in chunk_list(unique_models, MAX_MODELS_PER_RUN):
            item = self.run_actor(question, batch)
            model_responses = item.get("models", {})
            for model_name in batch:
                if model_name == PADDING_MODEL:
                    continue
                key = model_to_key(model_name)
                if key in model_responses:
                    answers[model_name] = str(model_responses[key])
        return answers


class AnthropicEvaluator:
    def __init__(self, api_key: str) -> None:
        self.client = Anthropic(api_key=api_key)

    def evaluate_answer(self, question: str, expected: str, answer: str) -> tuple[str, str]:
        prompt = build_evaluation_prompt(question, expected, answer)
        message = self.client.messages.create(
            model=ANTHROPIC_EVALUATOR_MODEL,
            max_tokens=256,
            messages=[{"role": "user", "content": prompt}],
        )
        text = "".join(block.text for block in message.content if block.type == "text")
        return parse_grade_and_comment(text)


def read_sheet_rows(ws: Worksheet, required_headers: list[str]) -> list[dict[str, str]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    header_index = {name: headers.index(name) for name in required_headers if name in headers}
    missing = [name for name in required_headers if name not in header_index]
    if missing:
        raise ValueError(f"Sheet '{ws.title}' is missing columns: {', '.join(missing)}")

    records: list[dict[str, str]] = []
    for row in rows[1:]:
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        record = {
            name: str(row[header_index[name]]).strip() if row[header_index[name]] is not None else ""
            for name in required_headers
        }
        if record[required_headers[0]]:
            records.append(record)
    return records


def normalize_model_header(headers: list[str]) -> str:
    for candidate in ("model", "model&version", "model_version", "model and version"):
        if candidate in headers:
            return candidate
    raise ValueError("Models sheet must include a 'model' or 'model&version' column")


def read_input_workbook(path: str) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    wb = load_workbook(path)
    if "questions" not in wb.sheetnames:
        raise ValueError("Workbook must contain a 'questions' sheet")
    if "models" not in wb.sheetnames:
        raise ValueError("Workbook must contain a 'models' sheet")

    questions = read_sheet_rows(wb["questions"], ["question", "answer_expected"])
    models_ws = wb["models"]
    model_rows = list(models_ws.iter_rows(values_only=True))
    if not model_rows:
        raise ValueError("Models sheet is empty")

    headers = [str(h).strip().lower() if h is not None else "" for h in model_rows[0]]
    model_col = normalize_model_header(headers)
    models = read_sheet_rows(models_ws, ["service_provider", model_col])
    for row in models:
        row["model"] = row.pop(model_col)
    return questions, models


def ensure_results_sheet(wb: Workbook) -> Worksheet:
    if "results" in wb.sheetnames:
        ws = wb["results"]
        existing = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        if existing != RESULT_COLUMNS:
            raise ValueError(
                f"Existing 'results' sheet headers must be: {RESULT_COLUMNS}. Found: {existing}"
            )
        return ws

    ws = wb.create_sheet("results")
    ws.append(RESULT_COLUMNS)
    return ws


def append_result(ws: Worksheet, row: dict[str, str]) -> None:
    ws.append([row[col] for col in RESULT_COLUMNS])


def create_template(path: str) -> None:
    wb = Workbook()
    questions_ws = wb.active
    questions_ws.title = "questions"
    questions_ws.append(["question", "answer_expected"])
    for question, expected in SAMPLE_QUESTIONS:
        questions_ws.append([question, expected])

    models_ws = wb.create_sheet("models")
    models_ws.append(["service_provider", "model&version"])
    for provider, model in DEFAULT_MODELS:
        models_ws.append([provider, model])

    results_ws = wb.create_sheet("results")
    results_ws.append(RESULT_COLUMNS)
    wb.save(path)


def run_comparison(input_path: str, output_path: str | None, delay_seconds: float) -> None:
    load_dotenv()
    apify_token = os.environ.get("APIFY_API_TOKEN", "").strip()
    if not apify_token:
        print("Error: set APIFY_API_TOKEN in your environment or .env file.", file=sys.stderr)
        sys.exit(1)

    anthropic_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if not anthropic_key:
        print("Error: set ANTHROPIC_API_KEY in your environment or .env file.", file=sys.stderr)
        sys.exit(1)

    questions, models = read_input_workbook(input_path)
    if not questions:
        raise ValueError("No questions found in the 'questions' sheet")
    if not models:
        raise ValueError("No models found in the 'models' sheet")

    for q in questions:
        if len(q["question"]) > MAX_PROMPT_CHARS:
            raise ValueError(
                f"Question exceeds {MAX_PROMPT_CHARS} characters (Apify limit): {q['question'][:80]}..."
            )

    runner = ModelComparisonRunner(apify_token)
    evaluator = AnthropicEvaluator(anthropic_key)
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    wb = load_workbook(input_path)
    results_ws = ensure_results_sheet(wb)
    save_path = output_path or input_path

    model_names = [m["model"] for m in models]
    provider_by_model = {m["model"]: m["service_provider"] for m in models}

    total = len(questions) * len(models)
    done = 0

    print(f"Run started at {timestamp}")
    print(f"Questions: {len(questions)}, models: {len(models)}, evaluator: {ANTHROPIC_EVALUATOR_MODEL}")

    for q in questions:
        question = q["question"]
        expected = q["answer_expected"]
        print(f"\nQuestion: {question}")

        answers = runner.get_model_answers(question, model_names)

        for model_name in model_names:
            answer = answers.get(model_name, "")
            if not answer:
                grade, comment = "", "No response returned from Apify actor"
            else:
                print(f"  Evaluating {model_name}...")
                grade, comment = evaluator.evaluate_answer(question, expected, answer)
                if delay_seconds:
                    time.sleep(delay_seconds)

            row = {
                "timestamp": timestamp,
                "question": question,
                "answer_expected": expected,
                "service_provider": provider_by_model.get(model_name, ""),
                "model": model_name,
                "answer": answer,
                "grade": grade,
                "comment": comment,
            }
            append_result(results_ws, row)
            wb.save(save_path)

            done += 1
            print(f"  [{done}/{total}] {model_name} -> grade {grade or 'n/a'}")

    print(f"\nDone. Results written to {save_path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Compare LLM model answers using Apify.")
    parser.add_argument(
        "input",
        nargs="?",
        default="comparison.xlsx",
        help="Input/output Excel file (default: comparison.xlsx)",
    )
    parser.add_argument(
        "-o",
        "--output",
        help="Optional output file (default: overwrite input file)",
    )
    parser.add_argument(
        "--create-template",
        action="store_true",
        help="Create a template Excel workbook and exit",
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=1.0,
        help="Seconds to wait between Anthropic grading calls (default: 1.0)",
    )
    args = parser.parse_args()

    if args.create_template:
        create_template(args.input)
        print(f"Template created: {args.input}")
        return

    run_comparison(args.input, args.output, args.delay)


if __name__ == "__main__":
    main()
